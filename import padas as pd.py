import pandas as pd
import os.path
from openpyxl import load_workbook
from operator import itemgetter
import json

class Excel_Parse: 
    def __init__(self, filename: str) -> str:
        self.workbook = filename
        self.file_contents = None
        self.sheet_names = None
        self.sheets = None
        
        self.read_csv()
        
    def read_csv(self):
        """
            loads the contents of an excel workbook into an object
        """
        if os.path.exists(self.workbook):
            self.file_contents = load_workbook(self.workbook, data_only=True)
            self.sheets = len(self.file_contents.sheetnames)
            self.sheetnames = self.file_contents.sheetnames
            
            return f"{filename} has been read successfully."
        else:
            return f"{filename} path does not exist."
        
    def get_sheet_names(self):
        """ 
            returns a list of sheet names in the excel workbook
        """
        return self.file_contents.sheetnames
        
    def get_sheet_tables(self, sheet_name: str) -> list:
        """
            sheet_name: str -> 
            returns a list of the tables 
        """
        try:
            ws = self.file_contents[sheet_name]
        except KeyError:
            return {"error": f"{sheet_name} does not exist"}    
        tables = [(sheet_name, key, value) for key, value in ws.tables.items()]
        return sorted(tables, key=lambda tables: tables[2])
        
        
    def get_table(self, table_range: tuple) -> dict:
        """
           table_range: tuple 
        
           returns the contents of one table in the requested table range as a dictionary.
          
        
        """
        data = self.file_contents[table_range[0]][table_range[2]]
        content = [[cell.value for cell in ent] for ent in data]
        header = content[0]
        rest = content[1:]
        df = pd.DataFrame(rest, columns=header)
        return json.loads(df.to_json(orient='records', indent=4))




FILENAME = r"FILENAME"                    # enter file name

# Some functions to complete proposed tasks

def return_all_sheets_all_tables():
    final = []
    excel = Excel_Parse(FILENAME)
    for name in excel.get_sheet_names():
        for tables in excel.get_sheet_tables(name):
            final.append(excel.get_table(tables))
    return json.dumps(final)


def return_one_sheet_all_tables():
    final = []
    excel = Excel_Parse(FILENAME)
    sheets = excel.get_sheet_names() 
    for tables in excel.get_sheet_tables(sheets[4]):
        final.append(excel.get_table(tables))
    return json.dumps(final)


def return_one_sheet_one_tables():
    final = []
    excel = Excel_Parse(FILENAME)
    sheets = excel.get_sheet_names() 
    final.append(excel.get_table(sheets[0])) 
    return json.dumps(final)